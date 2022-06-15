# encoding: utf-8
"""
Created on Fri Sep 11 19:11:20 2020

@author: Luiz
"""


import csv
from classes import Clientes, Agendamentos
from datetime import date,datetime,timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def getClientes():
    with open('clientes.csv', mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        clientes = []
        for row in csv_reader:
            clientes.append(Clientes(row['Nome'],row['Data de Nascimento'],
                                     row['CPF'],row['RG'],row['CEP'],row['Endereco'],
                                     row['Numero'],row['Complemento'],row['Bairro'],
                                     row['Cidade'],row['Estado'],row['DDD1'],
                                     row['Telefone1'],row['DDD2'],row['Telefone2'],
                                     row['Saldo'],row['Data'],row['ID']))
    return clientes

def saveClientes(clientes):
    with open('clientes.csv', mode='w', newline = '') as csv_file:
        fieldnames = ['Nome', 'Data de Nascimento','CPF','RG','CEP','Endereco','Numero','Complemento','Bairro','Cidade','Estado','DDD1','Telefone1','DDD2','Telefone2','Saldo','Data','ID']
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for cliente in clientes:
             writer.writerow({'Nome': cliente.nome,
                              'Data de Nascimento': cliente.ddn,
                              'CPF': cliente.cpf,
                              'RG':cliente.rg,
                              'CEP':cliente.cep,
                              'Endereco':cliente.endereco,
                              'Numero':cliente.numero,
                              'Complemento':cliente.complemento,
                              'Bairro':cliente.bairro,
                              'Cidade':cliente.cidade,
                              'Estado':cliente.estado,
                              'DDD1':cliente.ddd1,
                              'Telefone1':cliente.telefone1,
                              'DDD2':cliente.ddd2,
                              'Telefone2':cliente.telefone2,
                              'Saldo':cliente.saldo,
                              'Data':cliente.data,
                              'ID':cliente.ID})
 
def getAgendamentos():
    with open('agendamentos.csv', mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        agendamentos = []
        for row in csv_reader:
            agendamentos.append(Agendamentos(row['ID'],row['Evento'],row['Descricao'],
                                             row['Acessorios'],row['Ajustes'],row['Valor'],
                                             row['Desconto'],row['Sinal'],row['Data de Entrega'],
                                             row['Data de Devolucao']))
    
    return agendamentos

def getAgendamentosD():
    with open('agendamentosD.csv', mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        agendamentos = []
        for row in csv_reader:
            agendamentos.append(Agendamentos(row['ID'],row['Evento'],row['Descricao'],
                                             row['Acessorios'],row['Ajustes'],row['Valor'],
                                             row['Desconto'],row['Sinal'],row['Data de Entrega'],
                                             row['Data de Devolucao']))
    
    return agendamentos
       
def saveAgendamentos(agendamentos):
    with open('agendamentos.csv', mode='w', newline = '') as csv_file:
        fieldnames = ['ID', 'Evento','Descricao','Acessorios','Ajustes','Valor','Desconto','Sinal','Data de Entrega','Data de Devolucao']
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for agendamento in agendamentos:
             writer.writerow({'ID': agendamento.ID,
                              'Evento': agendamento.evento,
                              'Descricao': agendamento.descricao,
                              'Acessorios':agendamento.acessorios,
                              'Ajustes':agendamento.ajustes,
                              'Valor':agendamento.valor,
                              'Desconto':agendamento.desconto,
                              'Sinal':agendamento.sinal,
                              'Data de Entrega':agendamento.dde,
                              'Data de Devolucao':agendamento.ddd})             

def saveAgendamentosD(agendamentos):
    with open('agendamentosD.csv', mode='w', newline = '') as csv_file:
        fieldnames = ['ID', 'Evento','Descricao','Acessorios','Ajustes','Valor','Desconto','Sinal','Data de Entrega','Data de Devolucao']
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for agendamento in agendamentos:
             writer.writerow({'ID': agendamento.ID,
                              'Evento': agendamento.evento,
                              'Descricao': agendamento.descricao,
                              'Acessorios':agendamento.acessorios,
                              'Ajustes':agendamento.ajustes,
                              'Valor':agendamento.valor,
                              'Desconto':agendamento.desconto,
                              'Sinal':agendamento.sinal,
                              'Data de Entrega':agendamento.dde,
                              'Data de Devolucao':agendamento.ddd})  

def getCliente(Nome,CPF,clientes):
    for cliente in clientes:
        if((cliente.nome == Nome and Nome != "") or (cliente.cpf == CPF and CPF != "")):
            return cliente
    return "Not found"


def getAgendamento(ID,agendamentos):
    for agendamento in agendamentos:
        if(agendamento.ID == ID):
            return agendamento

    return "Not Found"


def getClientebyID(ID,clientes):
    for cliente in clientes:
        if((cliente.ID == ID and ID != "")):
            return cliente
    return "Not found"

def getSemanal():
    today = date.today()
    with open('agendamentos.csv', mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        agendamentos = []
        for row in csv_reader:
            agendamentos.append(Agendamentos(row['ID'],row['Evento'],row['Descricao'],
                                             row['Acessorios'],row['Ajustes'],row['Valor'],
                                             row['Desconto'],row['Sinal'],row['Data de Entrega'],
                                             row['Data de Devolucao']))
            
            
    
def saveEventos(lista):
 with open('eventos.txt', 'w', newline="") as txt_file:
     csv_writer = csv.writer(txt_file, delimiter = ',')
     for row in lista:
         if(row[1] >= date.today().strftime("%d/%m/%y") or row[1] >= date.today().strftime("%d/%m/%Y")):
             csv_writer.writerow(row)


def openEventos(evento):
    listas = getEventos()
    wb = Workbook()
    ws = wb.active
    value = 1
    for row in listas:
        ws.append(row)
        if(evento == row[0]):
            ws['A'+str(value)].fill = PatternFill("solid", fgColor="0000FF00")
        value+=1
    wb.save('eventos.xlsx')
    os.system("start EXCEL.EXE C:\\Users\\luiz_\\Desktop\\PalacioF\\programa\\interface\\eventos.xlsx")       
            
        
            

            
            
            
            